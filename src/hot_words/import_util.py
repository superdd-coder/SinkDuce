"""Parse hot-words CSV / Excel uploads into library fields."""

from __future__ import annotations

import csv
import io
import re
from pathlib import Path

from src.hot_words.models import HotWordItem

# Keep in sync with DashScope cloud language hints + frontend HOT_WORD_LANG_OPTIONS
ALLOWED_LANGS = frozenset({"", "zh", "en", "ja", "ko", "ms", "th", "id"})

_HEADER_ALIASES: dict[str, str] = {
    "text": "text",
    "word": "text",
    "hot_word": "text",
    "hotword": "text",
    "hotwords": "text",
    "term": "text",
    "phrase": "text",
    "weight": "weight",
    "w": "weight",
    "score": "weight",
    "priority": "weight",
    "lang": "lang",
    "language": "lang",
    "locale": "lang",
    "name": "name",
    "library": "name",
    "library_name": "name",
    "description": "description",
    "desc": "description",
}


def _norm_header(raw: str) -> str:
    s = (raw or "").strip().lower()
    s = s.replace("-", "_").replace(" ", "_")
    s = re.sub(r"_+", "_", s)
    return _HEADER_ALIASES.get(s, s)


def _clamp_weight(value: object, default: int = 4) -> int:
    if value is None or value == "":
        return default
    try:
        n = int(float(str(value).strip()))
    except (TypeError, ValueError):
        return default
    return max(1, min(10, n))


def _norm_lang(value: object) -> str:
    if value is None:
        return ""
    code = str(value).strip().lower()
    if code in ("auto", "any", "all", "*"):
        return ""
    if code not in ALLOWED_LANGS:
        # Keep free-text only if short ISO-like; otherwise drop
        if re.fullmatch(r"[a-z]{2,8}", code):
            return code
        return ""
    return code


def _row_dict(headers: list[str], cells: list[object]) -> dict[str, object]:
    out: dict[str, object] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        out[h] = cells[i] if i < len(cells) else ""
    return out


def rows_to_words(rows: list[dict[str, object]]) -> tuple[list[HotWordItem], str, str]:
    """Convert mapped row dicts → words + optional name/description from first non-empty meta."""
    words: list[HotWordItem] = []
    lib_name = ""
    lib_desc = ""
    for row in rows:
        if not lib_name:
            raw_name = row.get("name")
            if raw_name is not None and str(raw_name).strip():
                lib_name = str(raw_name).strip()
        if not lib_desc:
            raw_desc = row.get("description")
            if raw_desc is not None and str(raw_desc).strip():
                lib_desc = str(raw_desc).strip()

        text = str(row.get("text") or "").strip()
        if not text:
            continue
        words.append(
            HotWordItem(
                text=text,
                weight=_clamp_weight(row.get("weight")),
                lang=_norm_lang(row.get("lang")),
            )
        )
    return words, lib_name, lib_desc


def parse_csv_bytes(data: bytes) -> tuple[list[HotWordItem], str, str]:
    # Strip BOM
    text = data.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    rows_raw = list(reader)
    if not rows_raw:
        return [], "", ""

    headers = [_norm_header(c) for c in rows_raw[0]]
    # If no recognizable header, treat first col as text
    if "text" not in headers:
        # Heuristic: first row looks like data → invent headers
        if any(_norm_header(c) in _HEADER_ALIASES for c in rows_raw[0]):
            pass
        else:
            headers = ["text", "weight", "lang"][: len(rows_raw[0])]
            while len(headers) < 3:
                headers.append("")
            data_rows = rows_raw
            mapped = [_row_dict(headers, r) for r in data_rows]
            return rows_to_words(mapped)

    mapped = [_row_dict(headers, r) for r in rows_raw[1:]]
    return rows_to_words(mapped)


def parse_xlsx_bytes(data: bytes) -> tuple[list[HotWordItem], str, str]:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    try:
        ws = wb.active
        rows_raw: list[list[object]] = []
        for row in ws.iter_rows(values_only=True):
            cells = list(row)
            # Skip fully empty rows
            if all(c is None or str(c).strip() == "" for c in cells):
                continue
            rows_raw.append(cells)
    finally:
        wb.close()

    if not rows_raw:
        return [], "", ""

    headers = [_norm_header(str(c) if c is not None else "") for c in rows_raw[0]]
    if "text" not in headers:
        headers = ["text", "weight", "lang"]
        mapped = [_row_dict(headers, r) for r in rows_raw]
        return rows_to_words(mapped)

    mapped = [_row_dict(headers, r) for r in rows_raw[1:]]
    return rows_to_words(mapped)


def parse_hot_words_file(filename: str, data: bytes) -> tuple[list[HotWordItem], str, str]:
    """Return (words, name_from_file, description_from_file)."""
    suffix = Path(filename or "").suffix.lower()
    if suffix == ".csv":
        return parse_csv_bytes(data)
    if suffix in {".xlsx", ".xlsm"}:
        return parse_xlsx_bytes(data)
    if suffix == ".xls":
        raise ValueError(
            "Legacy .xls is not supported. Save as .xlsx or .csv and try again."
        )
    # Sniff by content
    if data[:2] == b"PK":
        return parse_xlsx_bytes(data)
    return parse_csv_bytes(data)


def build_template_csv() -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["text", "weight", "lang"])
    writer.writerow(["ExampleTerm", 4, "zh"])
    writer.writerow(["API", 5, "en"])
    writer.writerow(["", "", ""])
    # Comment-like guidance row is avoided (would import as a word); keep clean sample only
    return buf.getvalue().encode("utf-8-sig")


def build_template_xlsx() -> bytes:
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "hot_words"
    ws.append(["text", "weight", "lang"])
    ws.append(["ExampleTerm", 4, "zh"])
    ws.append(["API", 5, "en"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def build_export_xlsx(
    words: list[HotWordItem],
    *,
    name: str = "",
    description: str = "",
) -> bytes:
    """Export a library's words (and optional name/description columns) to .xlsx."""
    import openpyxl
    from openpyxl.styles import Font

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "hot_words"
    ws.append(["text", "weight", "lang", "name", "description"])
    for cell in ws[1]:
        cell.font = Font(bold=True)

    if not words:
        # Still write meta on an empty row so re-import keeps library name
        if name or description:
            ws.append(["", 4, "", name, description])
    else:
        for i, w in enumerate(words):
            ws.append(
                [
                    w.text,
                    w.weight,
                    w.lang or "",
                    name if i == 0 else "",
                    description if i == 0 else "",
                ]
            )

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 28
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
